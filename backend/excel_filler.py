"""
Excel template filler for Solar Load Analysis.
Creates a formatted Excel workbook matching the EnergyBae template.
"""

import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def fill_excel_template(data: dict) -> io.BytesIO:
    """
    Fill the solar load analysis Excel template with extracted bill data.
    Returns an in-memory BytesIO buffer with the Excel file.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Solar Load Analysis"

    # ── Styles ──
    header_font = Font(name='Calibri', bold=True, size=11)
    data_font = Font(name='Calibri', size=11)
    green_fill = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    orange_fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
    red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    light_blue_fill = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Set column widths
    col_widths = {'A': 8, 'B': 15, 'C': 18, 'D': 12, 'E': 14, 'F': 12, 'G': 15, 'H': 18, 'I': 12, 'J': 14, 'K': 12}
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

    # Extract data
    consumer_name = data.get('consumer_name', 'N/A')
    consumer_no = str(data.get('consumer_no', 'N/A'))
    fixed_charges = data.get('fixed_charges', 0)
    sanctioned_load = data.get('sanctioned_load_kw', 1.0)
    connection_type = data.get('connection_type', 'N/A')
    monthly_data = data.get('monthly_data', [])

    # ── Row 1: Consumer Name ──
    ws['B1'] = 'Consumer Name'
    ws['B1'].font = header_font
    ws['D1'] = consumer_name
    ws['D1'].font = Font(name='Calibri', bold=True, size=11, color='0000FF')
    ws['H1'] = consumer_name
    ws['H1'].font = Font(name='Calibri', bold=True, size=11, color='0000FF')

    # ── Row 2: Consumer No ──
    ws['B2'] = 'Consumer No'
    ws['B2'].font = header_font
    ws['D2'] = consumer_no
    ws['H2'] = consumer_no

    # ── Row 3: Fixed Charges ──
    ws['B3'] = 'Fixed Charges'
    ws['B3'].font = header_font
    ws['D3'] = fixed_charges
    ws['H3'] = fixed_charges

    # ── Row 4: Sanc. Load (KW) ──
    ws['B4'] = 'Sanc. Load (KW)'
    ws['B4'].font = header_font
    ws['D4'] = sanctioned_load
    ws['H4'] = sanctioned_load
    ws['D4'].fill = yellow_fill
    ws['H4'].fill = yellow_fill

    # ── Row 5: Connection Type ──
    ws['B5'] = 'Connection Type'
    ws['B5'].font = header_font
    ws['D5'] = connection_type
    ws['H5'] = connection_type

    # ── Row 6: Contract Demand (KVA) ──
    ws['B6'] = 'Contract Demand (KVA)'
    ws['B6'].font = header_font

    # ── Row 7: Solar Panel used ──
    ws['B7'] = 'Solar Panel used'
    ws['B7'].font = header_font
    ws['B7'].fill = green_fill
    ws['C7'] = 600
    ws['C7'].fill = green_fill

    # ── Row 8: Headers for monthly data ──
    headers_left = ['Sr.No', 'Month', 'Units', '', 'Bill Amount', 'Unit Cost']
    headers_right = ['Month', 'Units', '', 'Bill Amount', 'Unit Cost']
    cols_left = ['A', 'B', 'C', 'D', 'E', 'F']
    cols_right = ['G', 'H', 'I', 'J', 'K']

    for i, (col, header) in enumerate(zip(cols_left, headers_left)):
        cell = ws[f'{col}8']
        cell.value = header
        cell.font = header_font
        cell.fill = light_blue_fill
        cell.border = thin_border

    for i, (col, header) in enumerate(zip(cols_right, headers_right)):
        cell = ws[f'{col}8']
        cell.value = header
        cell.font = header_font
        cell.fill = light_blue_fill
        cell.border = thin_border

    # ── Rows 9-21: Monthly data (up to 13 months, split into two columns) ──
    # Left side: first set of data, Right side: second set if available
    half = len(monthly_data)

    for i, entry in enumerate(monthly_data):
        row = 9 + i
        sr_no = i + 1

        # Left side data
        ws[f'A{row}'] = sr_no
        ws[f'A{row}'].font = data_font
        ws[f'A{row}'].border = thin_border

        ws[f'B{row}'] = entry.get('month', '')
        ws[f'B{row}'].font = data_font
        ws[f'B{row}'].border = thin_border

        ws[f'C{row}'] = entry.get('units', 0)
        ws[f'C{row}'].font = data_font
        ws[f'C{row}'].border = thin_border

        if entry.get('bill_amount') is not None:
            ws[f'E{row}'] = entry['bill_amount']
            ws[f'E{row}'].font = data_font
            ws[f'E{row}'].border = thin_border

        if entry.get('unit_cost') is not None:
            ws[f'F{row}'] = entry['unit_cost']
            ws[f'F{row}'].font = data_font
            ws[f'F{row}'].border = thin_border

        # Right side (duplicate for the template format)
        ws[f'G{row}'] = entry.get('month', '')
        ws[f'G{row}'].font = data_font
        ws[f'G{row}'].border = thin_border

        ws[f'H{row}'] = entry.get('units', 0)
        ws[f'H{row}'].font = data_font
        ws[f'H{row}'].border = thin_border

    # ── Calculations Row (after monthly data) ──
    last_data_row = 9 + len(monthly_data) - 1
    calc_row = last_data_row + 2  # Leave a blank row

    # Calculate totals
    units_list = [e.get('units', 0) for e in monthly_data if e.get('units', 0) > 0]
    bill_amounts = [e.get('bill_amount', 0) for e in monthly_data if e.get('bill_amount')]
    unit_costs = [e.get('unit_cost', 0) for e in monthly_data if e.get('unit_cost')]

    avg_units = sum(units_list) / len(units_list) if units_list else 0
    avg_bill = sum(bill_amounts) / len(bill_amounts) if bill_amounts else 0
    avg_unit_cost = sum(unit_costs) / len(unit_costs) if unit_costs else 0

    # Average row
    ws[f'B{calc_row}'] = 'Average'
    ws[f'B{calc_row}'].font = header_font
    ws[f'C{calc_row}'] = round(avg_units, 2)
    ws[f'C{calc_row}'].font = header_font
    ws[f'E{calc_row}'] = round(avg_bill, 2) if avg_bill else ''
    ws[f'F{calc_row}'] = round(avg_unit_cost, 3) if avg_unit_cost else ''
    ws[f'G{calc_row}'] = 'Average'
    ws[f'G{calc_row}'].font = header_font
    ws[f'H{calc_row}'] = round(avg_units, 2)
    ws[f'J{calc_row}'] = round(avg_bill, 2) if avg_bill else ''
    ws[f'K{calc_row}'] = round(avg_unit_cost, 4) if avg_unit_cost else ''

    # KW row
    kw_row = calc_row + 1
    # KW = Average Units / (30 days * hours of sunlight ~5)
    kw_value = round(avg_units / 150, 2) if avg_units else 0
    ws[f'B{kw_row}'] = 'KW'
    ws[f'B{kw_row}'].font = header_font
    ws[f'C{kw_row}'] = round(kw_value, 8)
    ws[f'G{kw_row}'] = 'KW'
    ws[f'H{kw_row}'] = round(kw_value, 8)

    # Solar Panels row
    sp_row = kw_row + 1
    panel_watt = 600  # from row 7
    solar_panels_kw = round(kw_value * 1000 / panel_watt, 8) if kw_value else 0
    ws[f'B{sp_row}'] = 'Solar Panels'
    ws[f'B{sp_row}'].font = header_font
    ws[f'C{sp_row}'] = solar_panels_kw
    ws[f'G{sp_row}'] = 'Solar Panels'
    ws[f'H{sp_row}'] = solar_panels_kw

    # Solar capacity row
    sc_row = sp_row + 1
    import math
    solar_capacity = round(math.ceil(solar_panels_kw * 10) / 10, 1) if solar_panels_kw else 0
    ws[f'B{sc_row}'] = 'Solar capacity'
    ws[f'B{sc_row}'].font = header_font
    ws[f'B{sc_row}'].fill = orange_fill
    ws[f'C{sc_row}'] = solar_capacity
    ws[f'C{sc_row}'].fill = orange_fill
    ws[f'G{sc_row}'] = 'Solar capacity'
    ws[f'G{sc_row}'].font = header_font
    ws[f'G{sc_row}'].fill = orange_fill
    ws[f'H{sc_row}'] = solar_capacity
    ws[f'H{sc_row}'].fill = orange_fill

    # Number of Panels row
    np_row = sc_row + 1
    num_panels = math.ceil(solar_capacity * 1000 / panel_watt) if solar_capacity else 0
    ws[f'B{np_row}'] = 'Number of Panels'
    ws[f'B{np_row}'].font = header_font
    ws[f'B{np_row}'].fill = red_fill
    ws[f'B{np_row}'].font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
    ws[f'C{np_row}'] = num_panels
    ws[f'C{np_row}'].fill = red_fill
    ws[f'C{np_row}'].font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
    ws[f'G{np_row}'] = 'Number of Panels'
    ws[f'G{np_row}'].font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
    ws[f'G{np_row}'].fill = red_fill
    ws[f'H{np_row}'] = num_panels
    ws[f'H{np_row}'].fill = red_fill
    ws[f'H{np_row}'].font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')

    # Summary rows
    summary_row = np_row + 3
    ws[f'B{summary_row}'] = 'Total solar capacity'
    ws[f'B{summary_row}'].font = header_font
    ws[f'D{summary_row}'] = solar_capacity * 2  # Both sides combined
    ws[f'B{summary_row + 1}'] = 'Number of solar panels'
    ws[f'B{summary_row + 1}'].font = header_font
    ws[f'D{summary_row + 1}'] = num_panels * 2

    # Save to BytesIO
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
